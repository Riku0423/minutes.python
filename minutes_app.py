import os
import json
import google.generativeai as genai
import openpyxl
import logging
import argparse
from openpyxl.styles import Alignment
from openpyxl.utils.datetime import from_excel
from dotenv import load_dotenv
import subprocess
import concurrent.futures
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import sys
from pathlib import Path
import time
import google.api_core.exceptions
from docx import Document
import datetime
import xml.parsers.expat
from tkinter import ttk, font
from PIL import Image, ImageTk

# ユーザーディレクトリのDocumentsフォルダのパスを取得
documents_path = Path.home() / "Documents"
log_file_path = documents_path / "app_log.txt"

# ログファイルの設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path),  # ログファイルのパスを変更
        logging.StreamHandler()
    ]
)

def get_current_dir():
    if getattr(sys, 'frozen', False):
        # PyInstallerでパッケージ化された場合
        return Path(sys._MEIPASS)  # dist直下を基準にする
    else:
        # 開発環境の場合
        return Path(__file__).resolve().parent

# 現在のスクリプトのディレクトリを取得
current_dir = Path(__file__).resolve().parent

# 環境変数の読み込み
load_dotenv(current_dir / '環境変数.env')

# プロジェクトディレクトリの設定
project_dir = os.path.dirname(os.path.abspath(__file__))

# APIキーの設定
API_KEYS = [os.getenv(f'GEMINI_API_KEY_{i}') for i in range(1, 11)]  # 10個のAPIキーを取得

# 処理済みファイルのログファイル
PROCESSED_FILES_LOG = os.path.join(current_dir, 'processed_files.json')

def load_processed_files():
    if os.path.exists(PROCESSED_FILES_LOG):
        with open(PROCESSED_FILES_LOG, 'r') as f:
            return json.load(f)
    return {}

def save_processed_files(processed_files):
    with open(PROCESSED_FILES_LOG, 'w') as f:
        json.dump(processed_files, f, indent=2)

def get_unprocessed_audio_files():
    processed_files = load_processed_files()
    audio_files = [f for f in os.listdir(current_dir) if f.endswith('.mp3')]
    return [f for f in audio_files if f not in processed_files]

def create_extraction_prompt(text):
    return f"""
    この文章はとある会議の内容です。
    以下の文章から、次の項目を抽出してください：
    1. 議題①
    2. 議題①の要約
    3. 議題②
    4. 議題②の要約
    5. 議題③
    6. 議題③の要約
    7. 議題④
    8. 議題④の要約
    9. 議題⑤
    10. 議題⑤の要約
    11. 議題⑥
    12. 議題⑥の要約
    13. 議題⑦
    14. 議題⑦の要約
    15. 議題⑧
    16. 議題⑧の要約
    17. 議題⑨
    18. 議題⑨の要約
    19. 議題⑩
    20. 議題⑩の要約

    抽出する際は、必ず以下の形式で出力してください：
    議題①: [議題の内容]
    議題①の要約: [要約内容]

    議題②: [議題の内容]
    議題②の要約: [要約内容]

    議題③: [議題の内容]
    議題③の要約: [要約内容]

    ...

    議題⑩: [議題の内容]
    議題⑩の要約: [要約内容]

    注意事項:
    - 各議題とその要約を必ず上記の形式で出力してください。
    - 議題が10個未満の場合は、存在する議題のみを抽出してください。
    - 要約は簡潔かつ具体的にしてください。
    - 議題の番号（①、②など）は必ず付けてください。
    - 各行は必ず「議題○:」または「議題○の要約:」で始まるようにしてください。
    - 議題や要約の前に「*」や「**」などの記号を付けないでください。
    - 議題というのはあくまで表現の一つであり、会話内容が議事録形式で記されていれば構いません。インタビューの文章等からも適切に議題を抽出してください。
    - インタビューのような文章であっても、適切に議題を抽出してください。

    文章:
    {text}
    """


def get_ffmpeg_path():
    """ffmpegのパスを取得する関数"""
    # 環境変数からパスを取得
    ffmpeg_path = os.environ.get('FFMPEG_PATH')
    if ffmpeg_path:
        return ffmpeg_path

    # 実行ファイルと同じディレクトリにffmpegがあるか確認
    base_path = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    ffmpeg_path = base_path / 'ffmpeg.exe'
    if ffmpeg_path.exists():
        return str(ffmpeg_path)

    # システムのPATHから検索
    return 'ffmpeg'

def get_ffprobe_path():
    """ffprobeのパスを取得する関数"""
    # 環境変数からパスを取得
    ffprobe_path = os.environ.get('FFPROBE_PATH')
    if ffprobe_path:
        return ffprobe_path

    # 実行ファイルと同じディレクトリにffprobeがあるか確認
    base_path = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    ffprobe_path = base_path / 'ffprobe.exe'
    if ffprobe_path.exists():
        return str(ffprobe_path)

    # システムのPATHから検索
    return 'ffprobe'

def split_audio_file(audio_file_path, num_parts):
    """音声ファイルを指定された数の部分に重なりを持たせて分割する関数"""
    file_size = os.path.getsize(audio_file_path)
    duration = get_audio_duration(audio_file_path)  # 音声ファイルの長さを取得
    part_duration = duration / num_parts  # 各部分の長さ
    overlap_duration = part_duration * 0.1  # 10%の重なりを持たせる

    parts = []
    for i in range(num_parts):
        start_time = max(0, i * part_duration - (overlap_duration if i > 0 else 0))
        part_file = f"{audio_file_path}_part{i+1}.mp3"
        command = [
            str(get_ffmpeg_path()),  # ffmpegのパスを取得
            '-y',  # 追加: 出力ファイルが存在する場合は上書き
            '-i', audio_file_path,
            '-ss', str(start_time),
            '-t', str(part_duration + (overlap_duration if i < num_parts - 1 else 0)),
            '-c', 'copy',
            part_file
        ]
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding='utf-8')
        if result.returncode != 0:
            logging.error(f"FFmpegエラー: {result.stderr}")
        parts.append(part_file)

    return parts

def get_audio_duration(audio_file_path):
    """音声ファイルの長さを取得する関数"""
    command = [
        str(get_ffprobe_path()),  # ffprobeのパスを取得
        '-v', 'error',
        '-show_entries', 'format=duration',
        '-of', 'default=noprint_wrappers=1:nokey=1',
        audio_file_path
    ]
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    return float(result.stdout.strip())

def transcribe_audio_with_key(audio_file, api_key, retries=3):
    """指定されたAPIキーを使用して音声ファイルを文字起こしする���数"""
    for attempt in range(retries):
        try:
            with open(audio_file, 'rb') as audio:
                audio_data = audio.read()

            model = genai.GenerativeModel('gemini-1.5-pro')
            genai.configure(api_key=api_key)

            prompt = """
            以下の音声ファイルを文字起こししてください。以下の点に注意してください：
            1. 日本語で出力してください。
            2. 時間表記（例：13:05）は削除してください。
            3. 相槌（例：はい、うん、ええ）や言い淀み（例：あの、えーと）は削除してください。
            4. 文脈を損なわない範囲で、できるだけ簡潔に文字起こしを行ってください。
            5. 話者の区別は不要です。
            """

            response = model.generate_content(
                [
                    prompt,
                    {"mime_type": "audio/mp3", "data": audio_data}
                ]
            )

            if hasattr(response, 'text'):
                logging.info(f"{audio_file}の文字起こしが成功しました。")  # 成功メッセージのみ
                return response.text
            else:
                logging.error(f"文字起こし失敗: {audio_file} - レスポンスにテキストが含まれていません。")
        except google.api_core.exceptions.ResourceExhausted:
            logging.error(f"文字起こし失敗: {audio_file} - 429 Resource has been exhausted (e.g. check quota).")
        except Exception as e:
            logging.error(f"文字起こし失敗: {audio_file} - {str(e)}")
        
        if attempt < retries - 1:
            logging.info(f"リトライを試みます ({attempt + 2}/{retries})")
            time.sleep(60)
        else:
            logging.error(f"{audio_file}の文字起こしが{retries}回失敗しました。")
    
    return None

def extract_information(text):
    # logging.info(f"抽出前のテキスト: {text}")  # 抽出前のテキストをログに出力しない
    cleaned_text = " ".join(text.split())
    # logging.info(f"クリーンアップ後のテキスト: {cleaned_text}")  # クリーンアップ後のテキストをログに出力しない

    gemini_pro_api_key = os.getenv('GEMINI_API_KEY_1')
    if not gemini_pro_api_key:
        logging.error("Gemini-pro APIキーが設定されていません。環境変数.envファイルを確認してください。")
        return

    genai.configure(api_key=gemini_pro_api_key)
    model = genai.GenerativeModel('gemini-1.5-pro')
    
    # create_extraction_promptを使用してプロンプトを生成
    prompt = create_extraction_prompt(cleaned_text)

    try:
        logging.info("情報抽出を開始します。")
        response = model.generate_content(prompt)
        extracted_text = response.text.strip()
        logging.info(f"抽出結果全体: {extracted_text}")
        return extracted_text
    except Exception as e:
        logging.exception(f"情報抽出中にエラーが発生しました: {str(e)}")
        raise

def create_excel(extracted_info, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "議事録"

    # 列の幅を設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # 会議詳細情報を追加
    meeting_details = [
        "会議名",
        "日時",
        "場所",
        "参加者",
        "欠席者"
    ]

    for i, detail in enumerate(meeting_details, start=1):
        ws.cell(row=i, column=1, value=detail)
        ws.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=i, column=1).fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    row = 6  # 会議詳細情報の後から開始

    lines = extracted_info.split('\n')
    current_topic = ""
    current_summary = ""

    for line in lines:
        line = line.strip()
        if line.startswith("議題"):
            if current_topic and current_summary:
                # 前の議題を書き込む
                ws.cell(row=row, column=1, value=current_topic)
                cell = ws.cell(row=row, column=2, value=current_summary)
                cell.alignment = Alignment(wrap_text=True)
                row += 1
            parts = line.split(':', 1)
            if len(parts) == 2:
                current_topic = parts[0].strip()
                current_summary = parts[1].strip()
            else:
                current_topic = line
                current_summary = ""
        elif "の要約" in line:
            if current_topic and "の要約:" in line:
                current_summary = line.split("の要約:", 1)[1].strip()
        elif current_summary:
            current_summary += " " + line.strip()

    # 最後の議題を書き込む
    if current_topic and current_summary:
        ws.cell(row=row, column=1, value=current_topic)
        cell = ws.cell(row=row, column=2, value=current_summary)
        cell.alignment = Alignment(wrap_text=True)

    # セルのスタイルを設定
    for row in ws['A1:B'+str(ws.max_row)]:
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                                                 right=openpyxl.styles.Side(style='thin'), 
                                                 top=openpyxl.styles.Side(style='thin'), 
                                                 bottom=openpyxl.styles.Side(style='thin'))
            if cell.column == 1:  # A列のセルの場合
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            elif cell.column == 2:  # B列のセルの場合
                cell.alignment = Alignment(wrap_text=True)

    # B列の幅を内容に合わせて自動調整
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        if column_cells[0].column_letter == 'B':
            ws.column_dimensions[column_cells[0].column_letter].width = min(100, max(80, length))

    try:
        wb.save(output_file)
        logging.info(f"Excelファイルが正常に作成されました: {output_file}")
    except PermissionError:
        logging.error(f"Excelファイルの保存失敗しました。書き込み権限がありません: {output_file}")
    except Exception as e:
        logging.error(f"Excelファイルの保存中にエラーが発生しました: {str(e)}")

def process_audio_file(audio_file_path, processed_files):
    try:
        audio_file_name = os.path.basename(audio_file_path)
        file_size = os.path.getsize(audio_file_path)
        logging.info(f"{audio_file_name}の処理を開始します。ファイルサイズ: {file_size / (1024 * 1024):.2f}MB")

        # 音声ファイルを分割する数を決定
        num_parts = 10  # 10個のAPIキーを使用するため、分割数を10に設定

        transcribed_texts = [None] * num_parts  # インデックスに基づいて配置するリスト

        audio_parts = split_audio_file(audio_file_path, num_parts)

        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_to_index = {executor.submit(transcribe_audio_with_key, part, API_KEYS[i]): i for i, part in enumerate(audio_parts)}
            failed_parts = []
            for future in concurrent.futures.as_completed(future_to_index):
                index = future_to_index[future]
                part = audio_parts[index]
                result = future.result()
                if result:
                    transcribed_texts[index] = result
                    logging.info(f"{part}の処理が成功しました。")
                else:
                    logging.error(f"{part}の処理が失敗しました。")
                    failed_parts.append((index, part))

        # 失敗したパートのリトライ
        if failed_parts:
            logging.info("失敗したファイルのリトライを1分後に開始します。")
            time.sleep(60)
            for index, part in failed_parts:
                result = transcribe_audio_with_key(part, API_KEYS[0])
                if result:
                    transcribed_texts[index] = result
                    logging.info(f"{part}のリトライが成功しました。")
                else:
                    logging.error(f"{part}のリトライが失敗しました。")

        # 分割されたファイルを削除
        for part in audio_parts:
            os.remove(part)
        logging.info(f"{audio_file_name}の分割されたファイルを削除しました。")

        # 文字起こし結果を結合（Noneを除外）
        combined_text = "\n".join(filter(None, transcribed_texts))
        logging.info(f"{audio_file_name}の文字起こしが完了しました。情報を抽出します。")

        # 30秒のバッファを持たせる
        time.sleep(60)

        extracted_info = extract_information(combined_text)
        if extracted_info:
            output_file = os.path.join(Path.home(), 'Documents', f"{os.path.splitext(audio_file_name)[0]}_抽出結果.xlsx")
            create_excel(extracted_info, output_file)
            processed_files[audio_file_name] = output_file
        else:
            logging.error(f"{audio_file_name}の情報抽出に失敗しました。")

        return True
    except Exception as e:
        logging.exception(f"{audio_file_path}の処理中にエラーが発生しました: {str(e)}")
        return False

def extract_info_from_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {
        '会議名': sheet['B1'].value or '',
        '日時': convert_excel_date(sheet['B2'].value),
        '場所': sheet['B3'].value or '',
        '参加者': sheet['B4'].value or '',
        '欠席者': sheet['B5'].value or '',
    }
    for i in range(1, 11):  # 議題①から⑩まで
        data[f'議題{chr(0x2460 + i - 1)}'] = sheet[f'B{5+i*2-1}'].value or ''
        data[f'議題{chr(0x2460 + i - 1)}の要約'] = sheet[f'B{5+i*2}'].value or ''
    
    print("抽出されたデータ:")
    for key, value in data.items():
        print(f"{key}: {value}")
    
    return data

def convert_excel_date(value):
    if isinstance(value, (int, float)):
        return from_excel(value).strftime('%Y-%m-%d')
    return value

def create_minutes_from_template(data, template_path):
    # 修正後
    template_path = os.path.join(get_current_dir(), 'テンプレート.docx')

    doc = Document(template_path)
    
    for paragraph in doc.paragraphs:
        # 会議名、日時、場所、参加者、欠席者の置き換え
        for key, value in data.items():
            placeholder = f'「{key}」'
            if placeholder in paragraph.text:
                old_text = paragraph.text
                new_text = paragraph.text.replace(placeholder, str(value) if value is not None else '')
                paragraph.text = new_text
                print(f"置換: '{old_text}' -> '{new_text}'")

        # 議題と要約の置き換え
        for i in range(1, 11):
            topic_key = f'議題{chr(0x2460 + i - 1)}'
            topic_content = data.get(topic_key, '')
            summary_key = f'議題{chr(0x2460 + i - 1)}の要約'
            summary_content = data.get(summary_key, '')

            # 議題の名称を置き換え
            topic_placeholder = f'「{topic_key}」'
            if topic_placeholder in paragraph.text:
                old_text = paragraph.text
                new_text = paragraph.text.replace(topic_placeholder, topic_content)
                paragraph.text = new_text
                print(f"議題名置換: '{old_text}' -> '{new_text}'")

            # 要約の置き換え
            summary_placeholder = f'「{summary_key}」'
            if summary_placeholder in paragraph.text:
                old_text = paragraph.text
                new_text = paragraph.text.replace(summary_placeholder, summary_content)
                paragraph.text = new_text
                print(f"要約置換: '{old_text}' -> '{new_text}'")

    return doc

def create_minutes(xlsx_path, template_path, output_path):
    try:
        data = extract_info_from_xlsx(xlsx_path)
        doc = create_minutes_from_template(data, template_path)
        doc.save(output_path)
        print(f"議事録が作成されました: {output_path}")
        return True
    except Exception as e:
        logging.error(f"議事録の作成中にエラーが発生しました: {str(e)}")
        print(f"エラーが発生しました: {str(e)}")
        return False
    
# グローバル変数
selected_file = None
file_label = None
excel_file_label = None
uploading_label = None
elapsed_time_label = None
root = None

def show_main_menu():
    global root, file_label, excel_file_label, uploading_label, elapsed_time_label, selected_file
    selected_file = None
    for widget in root.winfo_children():
        widget.destroy()

    root.title("爆速議事録")
    root.geometry("1000x600")
    root.resizable(False, False)

    # 背景色を設定
    root.configure(bg="#f5f5f5")

    # タイトルラベル
    title_label = tk.Label(root, text="⚡️爆速議事録", font=("Noto Sans CJK JP", 40, "bold"), bg="#f5f5f5", fg="#333333")
    title_label.pack(pady=30)

    # メインフレーム
    main_frame = tk.Frame(root, bg="#f5f5f5")
    main_frame.pack(expand=True, fill="both", padx=50)

    # 音声ファイル処理フレーム
    audio_frame = create_process_frame(main_frame, "音声ファイル処理", upload_audio_file, complete_audio_upload)
    audio_frame.pack(side="left", padx=(0, 25))

    # Excelファイル処理フレーム
    excel_frame = create_process_frame(main_frame, "Excelファイル処理", upload_xlsx_file, complete_xlsx_upload)
    excel_frame.pack(side="right", padx=(25, 0))

def create_process_frame(parent, title, upload_func, process_func):
    frame = tk.Frame(parent, bg="white", bd=0, relief="ridge", width=400, height=450)
    frame.pack_propagate(False)

    title_label = tk.Label(frame, text=title, font=("Noto Sans CJK JP", 20, "bold"), bg="white", fg="#333333")
    title_label.pack(pady=20)

    upload_button = ttk.Button(frame, text="ファイルを選択", command=upload_func, style="TButton")
    upload_button.pack(pady=10)

    file_label = tk.Label(frame, text="選択したファイル", wraplength=350, justify="center", bg="white", fg="#666666", font=("Noto Sans CJK JP", 12))
    file_label.pack(pady=10)

    process_button = ttk.Button(frame, text="ファイルを処理", command=process_func, style="TButton")
    process_button.pack(pady=(20, 0))

    status_label = tk.Label(frame, text="", bg="white", fg="#666666", font=("Noto Sans CJK JP", 12))
    status_label.pack(pady=25)

    return frame

def configure_styles():
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("TButton", 
                    font=("Noto Sans CJK JP", 14),
                    background="#4CAF50",
                    foreground="white",
                    padding=10,
                    width=20)
    style.map("TButton",
              background=[('active', '#45a049')])

def upload_audio_file():
    global selected_file
    selected_file = filedialog.askopenfilename(filetypes=[("Audio Files", "*.wav *.mp3")])
    if selected_file:
        file_label.config(text=f"選択したファイル\n{os.path.basename(selected_file)}")

def complete_audio_upload():
    if selected_file:
        start_time = time.time()  # 処理開始時刻を記録
        uploading_label.config(text="音声ファイル処理中...")
        root.update_idletasks()
        processed_files = load_processed_files()
        threading.Thread(target=process_audio_file_async, args=(selected_file, processed_files, start_time)).start()
    else:
        messagebox.showwarning("警告", "ファイルが選択されていません。")

def upload_xlsx_file():
    global selected_file
    selected_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if selected_file:
        excel_file_label.config(text=f"選択したファイル\n{os.path.basename(selected_file)}")

def complete_xlsx_upload():
    if selected_file:
        elapsed_time_label.config(text="Excelファイル処理中...")
        root.update_idletasks()
        threading.Thread(target=process_xlsx_file_async, args=(selected_file,)).start()
    else:
        messagebox.showwarning("警告", "ファイルが選択されていません。")

def process_audio_file_async(audio_file, processed_files, start_time):
    def update_elapsed_time():
        while not processing_done:
            elapsed_time = int(time.time() - start_time)
            minutes, seconds = divmod(elapsed_time, 60)
            if minutes > 0:
                uploading_label.config(text=f"経過時間: {minutes}分{seconds}秒")
            else:
                uploading_label.config(text=f"経過時間: {seconds}秒")
            time.sleep(1)

    processing_done = False
    threading.Thread(target=update_elapsed_time).start()

    success = process_audio_file(audio_file, processed_files)
    processing_done = True
    total_elapsed_time = int(time.time() - start_time)
    minutes, seconds = divmod(total_elapsed_time, 60)
    if minutes > 0:
        uploading_label.config(text=f"処理にかかった時間: {minutes}分{seconds}秒で処理が完了しました")
    else:
        uploading_label.config(text=f"処理にかかった時間: {seconds}秒で処理が完了しました")

    if success:
        root.after(0, lambda: (messagebox.showinfo("完了", "ファイルのアップロードが完了しました。"), show_main_menu()))
    else:
        messagebox.showerror("エラー", "ファイルの処理中にエラーが発生しました。")

def process_xlsx_file_async(xlsx_file):
    template_path = os.path.join(get_current_dir(), 'テンプレート.docx')  # dist直下から取得
    output_path = os.path.join(Path.home(), 'Documents', f"{os.path.splitext(os.path.basename(xlsx_file))[0]}_議事録.docx")
    
    success = create_minutes(xlsx_file, template_path, output_path)
    elapsed_time_label.config(text="")
    if success:
        root.after(0, lambda: (messagebox.showinfo("完了", "議事録の作成が完了しました。"), show_main_menu()))
    else:
        messagebox.showerror("エラー", "ファイルの処理中にエラーが発生しました。")

def main():
    global root
    try:
        root = tk.Tk()
        # フォントの設定
        default_font = font.nametofont("TkDefaultFont")
        default_font.configure(family="Noto Sans CJK JP", size=12)
        root.option_add("*Font", default_font)
        
        configure_styles()
        show_main_menu()
        root.mainloop()
    except Exception as e:
        logging.exception("アプリケーションの実行中にエラーが発生しました。")
        messagebox.showerror("エラー", f"アプリケーションの実行中にエラーが発生しました:\n{str(e)}")
        logging.error(f"アプリケーションの起動時にエラーが発生しました: {str(e)}")

if __name__ == "__main__":
    main()
